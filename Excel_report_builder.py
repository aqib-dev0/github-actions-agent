import subprocess
import json
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime

def run_analysis():
    result = subprocess.run(["bash", "coverage_analysis.sh"], capture_output=True, text=True)
    return result.stdout

def create_excel(data):
    # Read existing CSV data
    historical_df = pd.read_csv('data.csv')
    
    # Get current year and month
    current_date = datetime.now()
    current_column = f"{current_date.year}-{current_date.month:02d}"
    
    # Check if current month already exists in CSV
    if current_column in historical_df.columns:
        # Get only last 4 months if current month exists
        historical_df = historical_df[['repo'] + historical_df.columns[-3:].tolist()]
    else:
        # Get only last 4 months
        historical_df = historical_df[['repo'] + historical_df.columns[-2:].tolist()]
        
        # Parse new JSON data
        coverage_data = json.loads(data)
        new_data = pd.DataFrame(list(coverage_data.items()), columns=['repo', current_column])
        
        # Merge historical data with new data
        historical_df = pd.merge(historical_df, new_data, on='repo', how='outer')
    
    # Create Excel file in memory
    excel_file = BytesIO()
    
    # Create Excel writer with openpyxl engine
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        historical_df.to_excel(writer, index=False, sheet_name='Coverage Report')
        
        # Get the worksheet
        worksheet = writer.sheets['Coverage Report']
        
        # Define light border style
        light_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        # Format header
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = light_border
            
        # Set fixed column width for better fit
        worksheet.column_dimensions['A'].width = 25  # Repo name column
        for col in range(2, worksheet.max_column + 1):
            worksheet.column_dimensions[worksheet.cell(1,col).column_letter].width = 12
            
        # Define fills for comparison
        increase_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')  # pale green
        decrease_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')  # pale red
        
        # Compare and color cells based on month-over-month changes
        for row in range(2, worksheet.max_row + 1):
            for col in range(3, worksheet.max_column + 1):  # Start from third column to compare with previous
                current_cell = worksheet.cell(row, col)
                prev_cell = worksheet.cell(row, col-1)
                
                # Add border to all cells
                current_cell.border = light_border
                prev_cell.border = light_border
                
                if current_cell.value and prev_cell.value:
                    current_val = float(str(current_cell.value).strip('%'))
                    prev_val = float(str(prev_cell.value).strip('%'))
                    
                    if current_val > prev_val:
                        current_cell.fill = increase_fill
                    elif current_val < prev_val:
                        current_cell.fill = decrease_fill
        
        # Center align all cells and ensure borders
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
                cell.border = light_border

    excel_file.seek(0)
    return excel_file, historical_df

def generate_file():
    analysis_results = run_analysis()
    excel_file, merged_df = create_excel(analysis_results)

    # Save updated data to CSV
    merged_df.to_csv('data.csv', index=False)
    return excel_file


